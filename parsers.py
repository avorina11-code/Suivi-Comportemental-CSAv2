# -*- coding: utf-8 -*-
"""
parsers.py
==========
Module de parsing pour l'application d'analyse d'adhérence planning CSA.

Ce module sait lire :
    1. Le Planning Hebdo (théorique)               -> load_planning()
    2. Les exports Vocalcom bruts (réel), qui existent sous 3 formes
       différentes selon ce que l'utilisateur exporte depuis Vocalcom :

       - "Agents pause report"        (OBLIGATOIRE) -> parse_pause_report()
         Grille horaire par tranche de 30 min (a.m./p.m.) + résumé
         (Arrivée-Départ, Durée de travail, Durée de pause).

       - "Agents report"              (optionnel)   -> parse_agents_report()
         Tableau détaillé par agent : Attente / Traitement / Post-travail /
         Pause avec Occurences, Temps total, Durée moyenne, etc.

       - "Agents activities by dates" (optionnel)   -> parse_activities_by_date()
         Résumé simplifié par agent : Arrivée-Départ + Durée de travail
         uniquement (les graphiques à barres empilées d'origine ne sont pas
         récupérables en tant que données lors de l'export Excel).

IMPORTANT — retour d'expérience sur des exports réels :
    Le nom de fichier donné par Vocalcom / l'utilisateur ne correspond PAS
    toujours à son contenu réel (ex: un fichier nommé
    "distribution_de_pause_agent.xls" peut en réalité contenir un export de
    type "Agents activities by dates", et inversement). Ce module NE FAIT
    JAMAIS confiance au nom du fichier : le type de rapport est détecté
    automatiquement en scannant le contenu des 100 premières lignes
    (fonction detect_report_type()). L'app affiche ce qui a été détecté
    pour chaque fichier importé.

IMPORTANT — format des fichiers Vocalcom :
    Les exports Vocalcom "Agents pause/activities report" sont très
    souvent au format binaire Excel 97-2003 (.xls, Composite Document /
    OLE), que la librairie moderne openpyxl NE SAIT PAS lire. Ce module
    passe donc par une grille normalisée (load_sheet_grid) qui utilise
    `xlrd` pour les .xls et `openpyxl` pour les .xlsx/.xlsm, afin que
    tout le reste du code de parsing soit indépendant du format d'origine.
"""

from __future__ import annotations

import io
import re
import datetime as dt
from dataclasses import dataclass
from typing import Optional

import pandas as pd
import numpy as np

# --------------------------------------------------------------------------
# Constantes
# --------------------------------------------------------------------------

JOURS_FR = ["Lundi", "Mardi", "Mercredi", "Jeudi", "Vendredi", "Samedi", "Dimanche"]

# Codes de statut "jour non travaillé" reconnus dans le Planning Hebdo
CODES_NON_TRAVAILLE = {"OFF", "CONGE", "CONGÉ", "PERMISSION", "FORMATION", "CM", "MALADIE"}

# Les 24 libellés de tranches de 30 min tels qu'ils apparaissent dans
# l'export "Agents pause report" (colonnes D à AA, index 3 à 26). Le même
# jeu de libellés sert pour la ligne "a.m." (00:00 -> 11:30) et la ligne
# "p.m." (12:00 -> 23:30).
SLOT_LABELS_RAW = [
    "12:00", "12:30", "01:00", "01:30", "02:00", "02:30", "03:00", "03:30",
    "04:00", "04:30", "05:00", "05:30", "06:00", "06:30", "07:00", "07:30",
    "08:00", "08:30", "09:00", "09:30", "10:00", "10:30", "11:00", "11:30",
]


def _slot_labels_24h(half: str) -> list[str]:
    """Convertit les 24 libellés bruts (affichage 12h) en libellés 24h réels
    selon que l'on est sur la ligne 'a.m.' (00:00-11:30) ou 'p.m.' (12:00-23:30).
    """
    out = []
    for i in range(len(SLOT_LABELS_RAW)):
        minutes_from_start = i * 30
        base_hour = 0 if half == "a.m." else 12
        total_min = base_hour * 60 + minutes_from_start
        h, m = divmod(total_min, 60)
        h = h % 24
        out.append(f"{h:02d}:{m:02d}")
    return out


SLOTS_AM_24H = _slot_labels_24h("a.m.")
SLOTS_PM_24H = _slot_labels_24h("p.m.")


# --------------------------------------------------------------------------
# Utilitaires génériques
# --------------------------------------------------------------------------

def clean_login(value) -> Optional[str]:
    """Nettoie un identifiant Login_Vocalcom : enlève espaces normaux et
    insécables (\\xa0, utilisés par Crystal Reports comme séparateur de
    milliers, ex: '1\\xa0055' -> '1055'), enlève un éventuel '.0' final
    laissé par une lecture Excel en float, et force en chaîne de caractères.
    """
    if value is None:
        return None
    s = str(value).strip()
    s = s.replace("\xa0", "").replace(" ", "")
    if s.endswith(".0"):
        s = s[:-2]
    digits = re.sub(r"\D", "", s)
    return digits if digits else None


def excel_fraction_to_time(value) -> Optional[dt.time]:
    """Convertit une fraction de jour Excel (ex: 0.291666..) en objet time.
    Les plannings Excel stockent souvent les heures comme des fractions de
    24h plutôt que comme des cellules 'Heure' formatées.
    """
    try:
        f = float(value)
    except (TypeError, ValueError):
        return None
    if f < 0 or f >= 1.5:  # >1 jour improbable pour une heure de shift
        return None
    total_seconds = round(f * 24 * 3600)
    h, rem = divmod(total_seconds, 3600)
    m, s = divmod(rem, 60)
    h = h % 24
    try:
        return dt.time(hour=h, minute=m, second=s)
    except ValueError:
        return None


def parse_hhmm_string(s: str) -> Optional[dt.time]:
    """Parse une heure au format '06h51', '06:51' ou '6h51' -> time(6,51)."""
    if s is None:
        return None
    s = str(s).strip()
    m = re.match(r"^(\d{1,2})[h:](\d{2})$", s)
    if not m:
        return None
    h, mi = int(m.group(1)), int(m.group(2))
    if h > 23 or mi > 59:
        return None
    return dt.time(hour=h, minute=mi)


def parse_duration_to_timedelta(s: str) -> Optional[dt.timedelta]:
    """Parse une durée Vocalcom vers un timedelta. Deux formats rencontrés :
       - "Xh MM'SS"  ex: "5h40'24"   -> 5h 40m 24s
       - "MM'SS"     ex: "03'47"     -> 0h 3m 47s (utilisé pour les cellules
         de tranche de 30 min)
    """
    if s is None:
        return None
    s = str(s).strip()
    m = re.match(r"^(\d+)h(\d{1,2})'(\d{1,2})$", s)
    if m:
        h, mi, se = (int(x) for x in m.groups())
        return dt.timedelta(hours=h, minutes=mi, seconds=se)
    m = re.match(r"^(\d{1,2})'(\d{1,2})$", s)
    if m:
        mi, se = (int(x) for x in m.groups())
        return dt.timedelta(minutes=mi, seconds=se)
    return None


def td_to_minutes(td: Optional[dt.timedelta]) -> float:
    if td is None:
        return 0.0
    return round(td.total_seconds() / 60.0, 2)


def date_to_jour_fr(d: dt.date) -> str:
    return JOURS_FR[d.weekday()]


# --------------------------------------------------------------------------
# Chargement bas niveau : grille normalisée (compatible .xls et .xlsx)
# --------------------------------------------------------------------------

def _read_bytes(file_like) -> bytes:
    if hasattr(file_like, "read"):
        data = file_like.read()
        if hasattr(file_like, "seek"):
            file_like.seek(0)
        return data
    if isinstance(file_like, str):
        with open(file_like, "rb") as fh:
            return fh.read()
    return file_like


def load_sheet_grid(file_like, filename: str = "") -> list[list]:
    """Charge la première feuille d'un classeur Excel (.xls ou .xlsx) et la
    retourne sous forme de grille normalisée : liste de lignes, chaque
    ligne étant une liste de valeurs (index 0 = colonne A, 1 = colonne B,
    etc.). Les cellules vides valent None. Toutes les lignes sont
    complétées (padding) à la même largeur pour simplifier les accès par
    index de colonne.

    - .xlsx / .xlsm -> openpyxl
    - .xls (binaire Excel 97-2003 / OLE, format très courant des exports
      Vocalcom/Crystal Reports) -> xlrd
    """
    data = _read_bytes(file_like)
    name_lower = (filename or getattr(file_like, "name", "") or "").lower()

    def _try_openpyxl():
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        ws = wb.worksheets[0]
        grid = [[c.value for c in row] for row in ws.iter_rows()]
        return grid

    def _try_xlrd():
        import xlrd
        wb = xlrd.open_workbook(file_contents=data)
        ws = wb.sheet_by_index(0)
        grid = [ws.row_values(r) for r in range(ws.nrows)]
        # xlrd renvoie '' pour les cellules vides -> on normalise en None
        for row in grid:
            for i, v in enumerate(row):
                if v == "":
                    row[i] = None
        return grid

    errors = []
    order = [_try_xlrd, _try_openpyxl] if name_lower.endswith(".xls") else [
        _try_openpyxl,
        _try_xlrd,
    ]
    for fn in order:
        try:
            grid = fn()
            width = max((len(r) for r in grid), default=0)
            return [r + [None] * (width - len(r)) for r in grid]
        except Exception as exc:  # noqa: BLE001
            errors.append(f"{fn.__name__}: {exc}")

    raise ValueError(
        "Impossible de lire ce fichier Excel (.xls/.xlsx). Formats testés "
        "sans succès (openpyxl et xlrd). Détails : " + " | ".join(errors)
    )


def _iter_cells_first_rows(grid, max_row=100):
    for row in grid[:max_row]:
        for v in row:
            if v is not None and str(v).strip() != "":
                yield v


# --------------------------------------------------------------------------
# 1. PLANNING HEBDO (théorique)
# --------------------------------------------------------------------------

def load_planning(file_like) -> pd.DataFrame:
    """Charge le fichier Planning Hebdo (.xlsx) -> DataFrame brut nettoyé.

    Colonnes en sortie : Matricule_RH, Login, Nom, puis pour chaque jour
    de JOURS_FR : '<Jour>_debut' et '<Jour>_fin', contenant soit un
    datetime.time, soit un code texte (OFF/CONGE/...), soit None.
    """
    df = pd.read_excel(file_like, sheet_name="Planning_Hebdo", engine="openpyxl")

    # Localisation souple des colonnes clés (au cas où l'intitulé varie légèrement)
    col_map = {}
    for col in df.columns:
        cl = str(col).strip().lower()
        if cl in ("login_vocalcom", "login", "id vocalcom"):
            col_map[col] = "Login"
        elif cl in ("matricule_rh", "matricule"):
            col_map[col] = "Matricule_RH"
        elif cl in ("nom_prenom", "nom", "nom_prénom"):
            col_map[col] = "Nom"
    df = df.rename(columns=col_map)

    if "Login" not in df.columns:
        raise ValueError(
            "Colonne Login_Vocalcom introuvable dans le Planning Hebdo. "
            "Vérifiez l'intitulé de la colonne."
        )

    df["Login"] = df["Login"].apply(clean_login)
    df = df.dropna(subset=["Login"]).reset_index(drop=True)

    def _convert(v):
        if v is None or (isinstance(v, float) and np.isnan(v)):
            return None
        if isinstance(v, (int, float)):
            return excel_fraction_to_time(v)
        if isinstance(v, dt.time):
            return v
        s = str(v).strip()
        if s.upper() in CODES_NON_TRAVAILLE:
            return s.upper()
        if s == "":
            return None
        # Shift éclaté "09:00/16:00" -> on garde la chaîne brute, elle
        # sera interprétée par get_shift_for_day().
        return s

    for jour in JOURS_FR:
        col_debut = f"{jour} Heure de début"
        col_fin = f"{jour} Heure de fin"
        df[f"{jour}_debut"] = df[col_debut].apply(_convert) if col_debut in df.columns else None
        df[f"{jour}_fin"] = df[col_fin].apply(_convert) if col_fin in df.columns else None

    keep_cols = ["Matricule_RH", "Login", "Nom"] + [
        f"{j}_{suf}" for j in JOURS_FR for suf in ("debut", "fin")
    ]
    keep_cols = [c for c in keep_cols if c in df.columns]
    return df[keep_cols].copy()


def _resolve_shift_bound(value, prendre_min: bool):
    """Résout une valeur de début/fin de shift (time, code statut, ou chaîne
    à tranches multiples séparées par '/') vers un objet time unique.
    Pour un shift éclaté, on prend la borne globale (début le plus tôt /
    fin la plus tardive) — voir limitation documentée dans le README.
    """
    if value is None:
        return None
    if isinstance(value, str) and value.upper() in CODES_NON_TRAVAILLE:
        return None
    if isinstance(value, dt.time):
        return value
    if isinstance(value, str) and "/" in value:
        parts = [parse_hhmm_string(p) or excel_fraction_to_time(p) for p in value.split("/")]
        parts = [p for p in parts if p is not None]
        if not parts:
            return None
        return min(parts) if prendre_min else max(parts)
    if isinstance(value, str):
        return parse_hhmm_string(value)
    return None


def get_shift_for_day(planning_df: pd.DataFrame, jour: str) -> pd.DataFrame:
    """Extrait, pour un jour de la semaine donné, le planning théorique de
    chaque agent -> DataFrame(Login, Matricule_RH, Nom, Debut_prevu,
    Fin_prevu, Statut_planning).

    Statut_planning ∈ {"TRAVAIL", "OFF", "CONGE", "PERMISSION",
    "FORMATION", "CM", "MALADIE", "NON_PLANIFIE"}.
    """
    col_debut = f"{jour}_debut"
    col_fin = f"{jour}_fin"
    rows = []
    for _, r in planning_df.iterrows():
        raw_debut = r.get(col_debut)
        raw_fin = r.get(col_fin)

        statut = "TRAVAIL"
        if raw_debut is None and raw_fin is None:
            statut = "NON_PLANIFIE"
        elif isinstance(raw_debut, str) and raw_debut.upper() in CODES_NON_TRAVAILLE:
            code = raw_debut.upper()
            statut = "CONGE" if code in ("CONGE", "CONGÉ") else code

        debut = _resolve_shift_bound(raw_debut, prendre_min=True) if statut == "TRAVAIL" else None
        fin = _resolve_shift_bound(raw_fin, prendre_min=False) if statut == "TRAVAIL" else None
        if statut == "TRAVAIL" and debut is None and fin is None:
            statut = "NON_PLANIFIE"

        rows.append(
            {
                "Login": r.get("Login"),
                "Matricule_RH": r.get("Matricule_RH"),
                "Nom": r.get("Nom"),
                "Debut_prevu": debut,
                "Fin_prevu": fin,
                "Statut_planning": statut,
            }
        )
    return pd.DataFrame(rows)


# --------------------------------------------------------------------------
# Détection automatique du type de rapport Vocalcom
# --------------------------------------------------------------------------

REPORT_TITLES = {
    "Agents pause report": "pause_report",
    "Agents report": "agents_report",
    "Agents activities by dates": "activities_by_date",
}


def detect_report_type(grid) -> str:
    """Scanne les 15 premières lignes pour déterminer le type réel du
    rapport, indépendamment du nom du fichier fourni par l'utilisateur
    (voir note en tête de module)."""
    for val in _iter_cells_first_rows(grid, max_row=15):
        s = str(val).strip()
        for title, key in REPORT_TITLES.items():
            if s == title:
                return key
    for val in _iter_cells_first_rows(grid, max_row=15):
        s = str(val).strip().lower()
        for title, key in REPORT_TITLES.items():
            if title.lower() in s:
                return key
    return "unknown"


def detect_report_date(grid) -> Optional[dt.date]:
    """Cherche un motif DD/MM/YYYY dans les 100 premières lignes."""
    pattern = re.compile(r"(\d{2})/(\d{2})/(\d{4})")
    for val in _iter_cells_first_rows(grid, max_row=100):
        m = pattern.search(str(val))
        if m:
            d, mo, y = (int(x) for x in m.groups())
            try:
                return dt.date(y, mo, d)
            except ValueError:
                continue
    return None


# --------------------------------------------------------------------------
# 2A. AGENTS PAUSE REPORT (obligatoire)
# --------------------------------------------------------------------------

AGENT_HEADER_RE = re.compile(r"^(\d[\d\s\xa0]*)\s*:\s*(.+)$")


@dataclass
class PauseReportResult:
    date: Optional[dt.date]
    summary: pd.DataFrame  # 1 ligne / agent
    timeline: pd.DataFrame  # 1 ligne / agent / tranche 30 min (pause uniquement)


def parse_pause_report(grid) -> PauseReportResult:
    """Parse la grille d'un export Vocalcom de type 'Agents pause report'.

    Pour chaque agent, on récupère :
      - Arrivée-Départ (première arrivée / dernier départ de la journée)
      - Durée de travail, Durée de pause, Trait. total (depuis le bloc
        'Résumé', seule source fiable pour les totaux)
      - La grille de consommation de pause par tranche de 30 minutes
        (uniquement les blocs libellés exactement 'Pause' ; d'autres
        catégories peuvent apparaître dans ce rapport, ex. 'Individual
        Coaching', mais elles sont ignorées pour la timeline de pause).
    """
    report_date = detect_report_date(grid)
    n = len(grid)

    summary_rows = []
    timeline_rows = []

    i = 0
    while i < n:
        row = grid[i]
        cell_a = row[0] if len(row) > 0 else None
        m = AGENT_HEADER_RE.match(str(cell_a).strip()) if cell_a is not None else None
        if not m:
            i += 1
            continue

        login = clean_login(m.group(1))
        nom = m.group(2).strip()

        pause_slots_am = [0.0] * 24
        pause_slots_pm = [0.0] * 24
        arrivee = depart = None
        duree_travail = duree_pause = trait_total = None

        j = i + 1
        current_label = None
        while j < n:
            r2 = grid[j]
            if all(v is None for v in r2):
                j += 1
                continue

            # Nouveau bloc agent -> on arrête le scan de celui-ci
            cell0 = r2[0]
            if cell0 is not None and AGENT_HEADER_RE.match(str(cell0).strip()):
                break

            col_b = r2[1] if len(r2) > 1 else None
            joined_b = str(col_b).strip() if col_b is not None else ""
            if joined_b == "Résumé":
                j += 1
                continue
            if joined_b.startswith("Trait. total"):
                trait_total = parse_duration_to_timedelta(joined_b.split(":", 1)[-1].strip())
                j += 1
                continue
            if joined_b.startswith("Durée de travail"):
                duree_travail = parse_duration_to_timedelta(joined_b.split(":", 1)[-1].strip())
                j += 1
                continue
            if joined_b.startswith("Durée de pause"):
                duree_pause = parse_duration_to_timedelta(joined_b.split(":", 1)[-1].strip())
                j += 1
                continue
            if joined_b.startswith("Arrivée-Départ"):
                plage = joined_b.split(":", 1)[-1].strip()
                if "-" in plage:
                    a, d = plage.split("-", 1)
                    arrivee = parse_hhmm_string(a.strip())
                    depart = parse_hhmm_string(d.strip())
                j += 1
                continue

            # Ligne d'en-tête de label d'activité (colonne A texte, non numérique)
            label_val = r2[0] if len(r2) > 0 else None
            if label_val is not None and not re.match(r"^\d", str(label_val).strip()):
                current_label = str(label_val).strip()

            # Ligne "a.m." / "p.m." avec valeurs de tranche (marqueur en col C, D ou E)
            marker_idx = None
            for idx in (2, 3, 4):
                if idx < len(r2) and str(r2[idx]).strip() in ("a.m.", "p.m."):
                    marker_idx = idx
                    break
            if marker_idx is not None and current_label == "Pause":
                half = str(r2[marker_idx]).strip()
                data_vals = r2[marker_idx + 1:]
                durations_min = []
                for v in data_vals:
                    if v is None:
                        continue
                    td = parse_duration_to_timedelta(str(v))
                    durations_min.append(td_to_minutes(td) if td else 0.0)
                if half == "a.m.":
                    # Quirk d'export connu : la 1ère tranche (00:00) est
                    # écrasée par le libellé 'a.m.' et n'est donc pas
                    # récupérable -> on la considère à 0, et on aligne les
                    # valeurs récupérées sur les tranches 1 à 23.
                    for k, v in enumerate(durations_min[:23]):
                        pause_slots_am[k + 1] += v
                else:
                    for k, v in enumerate(durations_min[:24]):
                        pause_slots_pm[k] += v
            j += 1

        summary_rows.append(
            {
                "Login": login,
                "Nom_Vocalcom": nom,
                "Arrivee_reelle": arrivee,
                "Depart_reel": depart,
                "Duree_travail_min": td_to_minutes(duree_travail),
                "Duree_pause_min": td_to_minutes(duree_pause),
                "Trait_total_min": td_to_minutes(trait_total),
            }
        )
        for k, label in enumerate(SLOTS_AM_24H):
            timeline_rows.append({"Login": login, "Slot": label, "Pause_min": pause_slots_am[k]})
        for k, label in enumerate(SLOTS_PM_24H):
            timeline_rows.append({"Login": login, "Slot": label, "Pause_min": pause_slots_pm[k]})

        i = j

    summary_df = pd.DataFrame(summary_rows)
    timeline_df = pd.DataFrame(timeline_rows)
    return PauseReportResult(date=report_date, summary=summary_df, timeline=timeline_df)


# --------------------------------------------------------------------------
# 2B. AGENTS REPORT (optionnel) -- détail Attente/Traitement/Pause/...
# --------------------------------------------------------------------------

AGENT_REPORT_HEADER_RE = re.compile(r"^Agent\s+(\d[\d\s\xa0]*)\s*:\s*(.+)$")

# Index de colonnes (0 = A) pour ce type de rapport
AR_COL_LABEL = 1     # B
AR_COL_OCC = 4        # E
AR_COL_TOTAL = 6      # G
AR_COL_PCT = 14       # O
AR_COL_SUBPCT = 16    # Q (présent uniquement sur les sous-catégories)


def parse_agents_report(grid) -> pd.DataFrame:
    """Parse la grille d'un export Vocalcom de type 'Agents report'
    (détail des états par agent : Attente, Traitement, Post-travail,
    Pause, etc.)

    -> DataFrame(Login, Nom, Temps_presence_min, Categorie, Occurences,
                 Temps_total_min, Pourcent_temps)

    Seules les catégories de premier niveau sont conservées ; les
    sous-catégories (ex: 'Appel entrant' sous 'Traitement', repérables à
    la présence d'un pourcentage en colonne Q) sont ignorées ici pour
    rester lisible.
    """
    rows = []
    current_login = current_nom = None
    current_presence = None
    presence_re = re.compile(r"(\d+h\d{1,2}'\d{1,2})")

    for r in grid:
        if all(v is None for v in r):
            continue

        a_val = r[0] if len(r) > 0 else None
        if a_val is not None:
            m = AGENT_REPORT_HEADER_RE.match(str(a_val).strip())
            if m:
                current_login = clean_login(m.group(1))
                current_nom = m.group(2).strip()
                presence_str = str(r[6]) if len(r) > 6 and r[6] is not None else ""
                pm = presence_re.search(presence_str)
                current_presence = td_to_minutes(
                    parse_duration_to_timedelta(pm.group(1)) if pm else None
                )
                continue

        label = r[AR_COL_LABEL] if len(r) > AR_COL_LABEL else None
        occ = r[AR_COL_OCC] if len(r) > AR_COL_OCC else None
        temps_total = r[AR_COL_TOTAL] if len(r) > AR_COL_TOTAL else None
        pourcent = r[AR_COL_PCT] if len(r) > AR_COL_PCT else None
        sub_pct = r[AR_COL_SUBPCT] if len(r) > AR_COL_SUBPCT else None

        if (
            current_login is not None
            and isinstance(label, str)
            and isinstance(occ, (int, float))
            and sub_pct is None
        ):
            rows.append(
                {
                    "Login": current_login,
                    "Nom": current_nom,
                    "Temps_presence_min": current_presence,
                    "Categorie": label.strip(),
                    "Occurences": int(occ),
                    "Temps_total_min": td_to_minutes(
                        parse_duration_to_timedelta(str(temps_total))
                    ),
                    "Pourcent_temps": pourcent,
                }
            )

    return pd.DataFrame(rows)


# --------------------------------------------------------------------------
# 2C. AGENTS ACTIVITIES BY DATES (optionnel) -- résumé simplifié
# --------------------------------------------------------------------------

ACTIVITIES_HEADER_RE = re.compile(r"^(\d[\d\s\xa0]*)\s*:\s*(.+)$")


def parse_activities_by_date(grid) -> pd.DataFrame:
    """Parse la grille d'un export Vocalcom de type 'Agents activities by
    dates'.

    ATTENTION : dans cet export, les graphiques à barres empilées
    d'origine (Attente/Supervision/Traitement/Post travail/Pause) ne sont
    PAS récupérables comme données une fois converti en fichier
    tabulaire — seules les informations Arrivée-Départ et Durée de
    travail totale restent exploitables. De plus le nom de l'agent y est
    tronqué (largeur de colonne fixe côté export) : on ne se fie donc
    qu'au Login pour la jointure, jamais au nom.

    -> DataFrame(Login, Nom_tronque, Arrivee, Depart, Duree_travail_min)
    """
    rows = []
    pending_plage = None  # 'HHhMM-HHhMM' vu juste avant la ligne agent
    plage_re = re.compile(r"^\d{2}h\d{2}-\d{2}h\d{2}$")

    for idx, r in enumerate(grid):
        if all(v is None for v in r):
            continue

        for v in r:
            if isinstance(v, str) and plage_re.match(v.strip()):
                pending_plage = v.strip()

        b_val = r[1] if len(r) > 1 else None
        if b_val is not None:
            m = ACTIVITIES_HEADER_RE.match(str(b_val).strip())
            if m:
                login = clean_login(m.group(1))
                nom_tronque = m.group(2).strip()
                duree = None
                if idx + 1 < len(grid):
                    for v in grid[idx + 1]:
                        if v is None:
                            continue
                        td = parse_duration_to_timedelta(str(v))
                        if td is not None:
                            duree = td
                            break
                arrivee = depart = None
                if pending_plage:
                    a, d = pending_plage.split("-")
                    arrivee = parse_hhmm_string(a)
                    depart = parse_hhmm_string(d)
                rows.append(
                    {
                        "Login": login,
                        "Nom_tronque": nom_tronque,
                        "Arrivee": arrivee,
                        "Depart": depart,
                        "Duree_travail_min": td_to_minutes(duree),
                    }
                )
                pending_plage = None

    return pd.DataFrame(rows)


# --------------------------------------------------------------------------
# Point d'entrée unique
# --------------------------------------------------------------------------

def load_any_vocalcom_export(file_like, filename: str = ""):
    """Détecte automatiquement le type d'export Vocalcom puis appelle le
    bon parseur. Retourne (type_detecte, resultat).

    type_detecte ∈ {"pause_report", "agents_report", "activities_by_date",
    "unknown"}.
    """
    grid = load_sheet_grid(file_like, filename=filename)
    rtype = detect_report_type(grid)

    if rtype == "pause_report":
        return rtype, parse_pause_report(grid)
    if rtype == "agents_report":
        return rtype, parse_agents_report(grid)
    if rtype == "activities_by_date":
        return rtype, parse_activities_by_date(grid)
    return rtype, None
